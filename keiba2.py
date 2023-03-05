#!/usr/bin/env python
# coding: utf-8

import os
import time
import requests
import pandas as pd
import datetime
import collections
import openpyxl
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor
from itertools import repeat
from openpyxl.styles.alignment import Alignment

number_of_max_workers=None
waiting_to_load_sec = 1
time_to_unload_sec = 60
retry_count = 3
domain = "https://race.netkeiba.com/race/shutuba_past.html?race_id="

# ====================
# 編集エリア start
# ====================

dir_name = 'test' # フォルダ名
file_name = 'output.xlsx' # エクセルファイル名
# dir_name = 'result' # フォルダ名
# file_name = 'output.xlsx' # エクセルファイル名
sheet_title = '出走機会' # エクセルシート名
output_location = '.' # ファイルの出力先

# 取得したい年
from_year = 2020
to_year = 2023

# 開催数(第○回)
# 最大 = '05'
# '01', '02', '03', '04', '05'
times_code_list = ['03', '04', '05']

# 開催日(○日目)
# 最大 = '09'
# '01', '02', '03', '04', '05', '06', '07', '08', '09'
date_code_list = ['01', '02', '03', '04', '05', '06', '07', '08', '09']

# 開催地コード
# '01': '札幌', '02': '函館', '03': '福島', '04': '新潟', '05': '東京', '06': '中山', '07': '中京', '08': '京都', '09': '阪神', '10': '小倉'
venue_code_list = {'01': '札幌'}

# 1日のレース数
# 最大 = '12'
# '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12'
race_code_list = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']

# ====================
# 編集エリア end
# ====================

# create soup
def create_soup(url):
    for i in range(retry_count):
        try:
            result = requests.get(url)
            time.sleep(waiting_to_load_sec)
            if result.status_code == 200:
                break
            else:
                print(f'{url} Retry count: {i+1}')
                time.sleep(waiting_to_load_sec)
        except requests.exceptions.HTTPError as err:
            print(f'err: {err}')
            print(f'url: {url}')

    soup = BeautifulSoup(result.content, 'html.parser')
    time.sleep(waiting_to_load_sec)
    return soup

# 日時判定(リクエストレースが未来日付の場合処理しない)
def is_date(soup, year):
    date_element = soup.find('dd', class_="Active")
    today = datetime.datetime.today()
    if date_element == None:
        return False
    date = date_element.find('a').text
    if date.find('/') != -1 and date.find('(') == -1:
        screen_date = date.replace('/', '-')
    elif date.find('/') == -1 and date.find('(') != -1:
        screen_date = date_element.find('a').text[:date_element.find('a').text.find('日')].replace('月', '-')
    else:
        return False
    return today > datetime.datetime.strptime(f'{year}-{screen_date} 23:59:59', '%Y-%m-%d %H:%M:%S')

# レースデータ取得
def get_race_data_list(soup):
    race_info_contents = soup.find('div', class_="RaceList_Item02")
    race_data = race_info_contents.find('div', class_="RaceData01")
    return race_data.find_all('span')

# レースの馬場取得
def get_race_track(race_data_list):
    track = race_data_list[0].string.strip()[0] if bool(race_data_list[0].string) else 'unknown'
    return "ダート" if track == "ダ" else track

# データ取得し振分(ダート or 芝)
def create_data_result(domain, target_year, venue_code, times_code, date_code, dirt_race_list, turf_race_list, race_code):
    url = f'{domain}{target_year}{venue_code}{times_code}{date_code}{race_code}'
    soup = create_soup(url)
    horse_table_contents = soup.find('table', class_="Shutuba_Table Shutuba_Past5_Table")
    get_horse_detail_tr = horse_table_contents.find_all("tr")
    date = is_date(soup, target_year)
    if not bool(date):
        # レースが存在しない or まだ行われていない
        time.sleep(waiting_to_load_sec * 3)
        return

    race_data_list = get_race_data_list(soup)
    track = get_race_track(race_data_list)
    for i in range(len(get_horse_detail_tr)):
        if i == 0:
            continue
        name = get_horse_detail_tr[i].find('div', class_="Horse01 fc")
        if track == 'ダート':
            dirt_race_list.append(name.string)
        elif track == '芝':
            turf_race_list.append(name.string)
        else:
            continue

# データ集計
def format_data_result(dirt_race_list, turf_race_list):
    dirt_race_count = collections.Counter(dirt_race_list)
    turf_race_count = collections.Counter(turf_race_list)
    print(f'dirt_race_list: {dirt_race_list}')
    print(f'turf_race_list: {turf_race_list}')
    print(f'dirt_race_count: {dirt_race_count}')
    print(f'turf_race_count: {turf_race_count}')
    result = []
    for name in dirt_race_count:
        horse_data = {}
        horse_data['name'] = name
        horse_data['dirt'] = dirt_race_count[name]
        horse_data['turf'] = turf_race_count[name]
        result.append(horse_data)

    for name in turf_race_list:
        # 出走機会が芝のみの場合を考慮
        if not name in dirt_race_count:
            horse_data = {}
            horse_data['name'] = name
            horse_data['dirt'] = dirt_race_count[name]
            horse_data['turf'] = turf_race_count[name]
            result.append(horse_data)
    print(f'result: {result}')
    return result

# エクセルファイルの作成
def create_excel_file():
    if not os.path.exists(f'{output_location}/{dir_name}'):
        # ディレクトリの存在確認
        os.mkdir(f'{output_location}/{dir_name}')

    if os.path.exists(f'{output_location}/{dir_name}/{file_name}'):
        # ファイルが存在する場合読み込み
        wb = openpyxl.load_workbook(f'{output_location}/{dir_name}/{file_name}')
        return wb
    else:
        # ファイルが存在しない場合新規作成
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_title
        wb.save(f'{output_location}/{dir_name}/{file_name}')
        return wb

# エクセルファイルの更新
def update_excel_file(wb, index, venue_code, result):
    row_num_1 = 1
    row_num_2 = 2
    row_num_3 = 3
    col_num_1 = 1
    col_num_2 = 2
    col_num_3 = 3
    horse_name_cell_width = 25
    ws = wb[sheet_title]
    # セルを結合し中央揃え(会場)
    ws.merge_cells(
        start_row = row_num_2,
        end_row = row_num_2,
        start_column = index + col_num_2 + index,
        end_column = index + col_num_3 + index
    )
    ws.cell(row_num_2, (index + col_num_2 + index), value=venue_code_list[venue_code])
    ws.cell(row_num_2, (index + col_num_2 + index)).alignment = Alignment(horizontal='center')
    # セルを作成し中央揃え
    ws.cell(row_num_3, (index + col_num_2 + index), value="芝")
    ws.cell(row_num_3, (index + col_num_2 + index)).alignment = Alignment(horizontal='center')
    ws.cell(row_num_3, (index + col_num_3 + index), value="ダート")
    ws.cell(row_num_3, (index + col_num_3 + index)).alignment = Alignment(horizontal='center')
    # 書き込み
    for i, horse in enumerate(result):
        if index == 0:
            # ファイル作成初回時
            ws.cell(row_num_1, col_num_1, value=f'{from_year}-{to_year}') # A1列
            ws.cell(row_num_3, col_num_1, value="種牡馬") # A3列
            ws.cell(row_num_3, col_num_1).alignment = Alignment(horizontal='center')
            ws.column_dimensions['A'].width = horse_name_cell_width
            ws.cell(i + 1 + row_num_3, col_num_1, value=horse['name']) # 馬名のセル(A列)
            ws.cell(i + 1 + row_num_3, col_num_2, value=horse['turf']) # 芝の出走機会
            ws.cell(i + 1 + row_num_3, col_num_3, value=horse['dirt']) # ダートの出走機会
        else:
            # ファイル作成2回目以降
            maxRow = ws.max_row + 1
            name_result = []
            for col in ws['A']:
                value = str(col.value)
                name_result.append(value)
            if not horse['name'] in name_result:
                # A列に名前がない場合新規で列作成し書き込み
                ws.cell(maxRow, col_num_1, value=horse['name']) # 馬名のセル(A列)
                ws.cell(maxRow, index + col_num_2 + index, value=horse['turf']) # 芝の出走機会
                ws.cell(maxRow, index + col_num_3 + index, value=horse['dirt']) # ダートの出走機会
            else:
                # A列に名前がある場合 target_row を取得し書き込み
                target_row = name_result.index(horse['name']) + 1
                ws.cell(target_row, index + col_num_2 + index, value=horse['turf']) # 芝の出走機会
                ws.cell(target_row, index + col_num_3 + index, value=horse['dirt']) # ダートの出走機会
        wb.save(f'{output_location}/{dir_name}/{file_name}')
    print('')
    print(f'[{get_now_time()}] ◇◇◇ Successful file saved ◇◇◇')
    print(f'[{get_now_time()}] ◇◇◇ Output location is {output_location}/{dir_name}/{file_name} ◇◇◇')
    print('')

# 現在時刻取得(ログ出力)
def get_now_time():
    dt_now = datetime.datetime.now()
    return dt_now.strftime('%Y-%m-%d %H:%M:%S')

for index, venue_code in enumerate(venue_code_list):
    dirt_race_list = []
    turf_race_list = []
    print('========================================')
    print(f'[{get_now_time()}] Start {venue_code_list[venue_code]} Race')
    print('========================================')

    for year in range(to_year - from_year + 1):
        target_year = from_year + year
        print('')
        print(f'[{get_now_time()}] Starting {target_year} {venue_code_list[venue_code]} 🏇🏇🏇')
        print('')

        for times_code in times_code_list:
            if times_code == '01':
                suffix = 'st'
            elif times_code == '02':
                suffix = 'nd'
            elif times_code == '03':
                suffix = 'rd'
            else:
                suffix = 'th'
            print(f'[{get_now_time()}] {venue_code_list[venue_code]} {times_code}{suffix} ===>>>')

            for date_code in date_code_list:
                print(f'[{get_now_time()}] Day {date_code} Loading...')
                with ThreadPoolExecutor(max_workers=number_of_max_workers) as executor:
                    executor.map(create_data_result, repeat(domain), repeat(target_year), repeat(venue_code), repeat(times_code), repeat(date_code), repeat(dirt_race_list), repeat(turf_race_list), race_code_list)

            print(f'[{get_now_time()}] {venue_code_list[venue_code]} {times_code}{suffix} <<<===')

        print('')
        print(f'[{get_now_time()}] Finished {target_year} {venue_code_list[venue_code]}')
        print(f'[{get_now_time()}] Wait for {time_to_unload_sec}sec.')
        time.sleep(time_to_unload_sec)

    result = format_data_result(dirt_race_list, turf_race_list)
    wb = create_excel_file()
    update_excel_file(wb, index, venue_code, result)
    print('========================================')
    print(f'[{get_now_time()}] Finished {venue_code_list[venue_code]} Race')
    print('========================================')
    print('')
    print(f'[{get_now_time()}] Wait for {time_to_unload_sec * 3}sec.')
    time.sleep(time_to_unload_sec * 3)

print('========================================')
print(f'[{get_now_time()}] All work done')
print('========================================')
print('')
